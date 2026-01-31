import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import io
import re
import random

# ==========================================
# 🔒 安全守門員：登入檢查系統
# ==========================================
def check_password():
    """如果使用者輸入正確密碼，回傳 True，否則回傳 False"""

    def password_entered():
        """檢查使用者輸入的密碼是否與 secrets 中的設定相符"""
        if st.session_state["password"] == st.secrets["LOGIN_PASSWORD"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # 驗證後刪除輸入框的暫存，保持乾淨
        else:
            st.session_state["password_correct"] = False

    # 初始化 session state
    if "password_correct" not in st.session_state:
        # 第一次進入，顯示輸入框
        st.text_input(
            "請輸入系統密碼 / Password", type="password", on_change=password_entered, key="password"
        )
        return False
    
    elif not st.session_state["password_correct"]:
        # 密碼錯誤，再次顯示輸入框
        st.text_input(
            "❌ 密碼錯誤，請重試 / Password", type="password", on_change=password_entered, key="password"
        )
        return False
    
    else:
        # 密碼正確
        return True

# 🚨 執行檢查：如果沒通過，程式就停在這裡 (st.stop)
if not check_password():
    st.stop()

# ==========================================
# 👇 只有登入成功後，才會執行下面的程式碼
# ==========================================

# ==========================================
# ⚙️ 第一部分：產生模板邏輯 (V5 + 真實資料預填)
# ==========================================
def generate_template_bytes(year, month):
    wb = Workbook()
    
    # 樣式定義
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Sheet 0: 全域控制台
    ws0 = wb.active; ws0.title = "0_全域控制台"
    ws0.append(['項目', '數值', '說明'])
    ws0.append(['年份', year, '設定排班年份'])
    ws0.append(['月份', month, '設定排班月份'])
    for cell in ws0[1]: cell.font = font_header; cell.fill = fill_header

    # Sheet 1: 行事曆與醫師 (更新：加入真實醫師預填)
    ws1 = wb.create_sheet("1_行事曆與醫師")
    dates = pd.date_range(start=f'{year}-{month}-01', end=f'{year}-{month}-{pd.Period(f"{year}-{month}").days_in_month}')
    weekday_map = {0:'一', 1:'二', 2:'三', 3:'四', 4:'五', 5:'六', 6:'日'}
    headers1 = ['日期', '星期', '時段', '丁院_醫師', '戊院_醫師', '丁_PT需求', '丁_OT需求', '戊_PT需求', '戊_OT需求', '營業狀態']
    ws1.append(headers1)
    
    # ★★★ 真實醫師班表設定 (依照圖一) ★★★
    # 格式: '班別': {'d_doc': 丁醫, 'w_doc': 戊醫, 'd_pt': 丁P, 'd_ot': 丁O, 'w_pt': 戊P, 'w_ot': 戊O}
    WEEKLY_TEMPLATE = {
        0: { # 週一
            'A': {'d_doc': '劉醫師', 'w_doc': '薛醫師', 'd_pt': 5, 'd_ot': 0, 'w_pt': 4, 'w_ot': 0},
            'B': {'d_doc': '莊醫師', 'w_doc': '劉醫師', 'd_pt': 4, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
            'C': {'d_doc': '劉醫師', 'w_doc': '莊醫師', 'd_pt': 4, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
        },
        1: { # 週二
            'A': {'d_doc': '莊醫師', 'w_doc': '薛醫師', 'd_pt': 4, 'd_ot': 0, 'w_pt': 4, 'w_ot': 0},
            'B': {'d_doc': '劉醫師', 'w_doc': '王醫師', 'd_pt': 4, 'd_ot': 0, 'w_pt': 4, 'w_ot': 0},
            'C': {'d_doc': '薛醫師', 'w_doc': '王醫師', 'd_pt': 3, 'd_ot': 0, 'w_pt': 4, 'w_ot': 0},
        },
        2: { # 週三
            'A': {'d_doc': '薛醫師', 'w_doc': '劉醫師', 'd_pt': 4, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
            'B': {'d_doc': '莊醫師', 'w_doc': '王醫師', 'd_pt': 3, 'd_ot': 0, 'w_pt': 4, 'w_ot': 0},
            'C': {'d_doc': '王醫師', 'w_doc': '莊醫師', 'd_pt': 3, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
        },
        3: { # 週四
            'A': {'d_doc': '莊醫師', 'w_doc': '劉醫師', 'd_pt': 4, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
            'B': {'d_doc': '王醫師', 'w_doc': '無',     'd_pt': 4, 'd_ot': 0, 'w_pt': 3, 'w_ot': 0},
            'C': {'d_doc': '王醫師', 'w_doc': '劉醫師', 'd_pt': 4, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
        },
        4: { # 週五
            'A': {'d_doc': '劉醫師', 'w_doc': '薛醫師', 'd_pt': 5, 'd_ot': 0, 'w_pt': 4, 'w_ot': 0},
            'B': {'d_doc': '無',     'w_doc': '莊醫師', 'd_pt': 3, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
            'C': {'d_doc': '莊醫師', 'w_doc': '劉醫師', 'd_pt': 3, 'd_ot': 0, 'w_pt': 3, 'w_ot': 1},
        }
    }

    row_count = 1
    for d in dates:
        if d.weekday() >= 5: continue # 跳過六日
        d_str = d.strftime('%Y/%m/%d')
        wk = weekday_map[d.weekday()]
        daily_plan = WEEKLY_TEMPLATE.get(d.weekday(), {})
        
        for shift in ['A', 'B', 'C']:
            sp = daily_plan.get(shift, {})
            ws1.append([
                d_str, wk, shift, 
                sp.get('d_doc',''), sp.get('w_doc',''), 
                sp.get('d_pt',3), sp.get('d_ot',0), 
                sp.get('w_pt',3), sp.get('w_ot',0), 
                '營業'
            ])
            row_count += 1
            
    for cell in ws1[1]: cell.font = font_header; cell.fill = fill_header; cell.alignment = center_align
    if row_count > 1:
        dv = DataValidation(type="list", formula1='"營業,休診"', allow_blank=False)
        ws1.add_data_validation(dv); dv.add(f'J2:J{row_count}')

    # Sheet 2: 人員設定 (更新：填入圖二真實名單)
    ws2 = wb.create_sheet("2_人員設定")
    headers2 = ['序號', '姓名', '員工編號', '身分 (下拉)', '職能 (下拉)', '本月目標診數', '備註', '週一 (固定/可排)', '週二 (固定/可排)', '週三 (固定/可排)', '週四 (固定/可排)', '週五 (固定/可排)']
    ws2.append(headers2)
    
    real_staff_data = [
        [1, '林振明', 'PTA005', 'FT', 'PT(物治)', 40, '', '', '', '', 'A甲', ''],
        [2, '張雅惠', 'A002', 'FT', 'PT(物治)', 40, '', '', 'B甲', 'A甲', '', ''],
        [3, '曾詩婷', 'PT022', 'FT', 'PT(物治)', 40, '', '', 'C甲', 'C甲', '', ''],
        [4, '葉宜甫', 'PT037', 'FT', 'PT(物治)', 40, '', '', '', '', 'C甲', ''],
        [5, '吳星霈', 'PT044', 'FT', 'PT(物治)', 40, '', 'B甲', '', '', '', ''],
        [6, '廖姿雅', 'PT031', 'FT', 'PT(物治)', 40, '', 'C甲', '', '', '', ''],
        [7, '林艾炘', 'PT043', 'FT', 'PT(物治)', 40, '', '', '', '', '', 'B甲'],
        [8, '鄭詠心', 'PTP116', 'FT', 'PT(物治)', 40, '', '', '', '', '', 'C甲'],
        [9, '鄧雅曼', 'OT022', 'FT', 'OT(職治)', 40, '', 'B戊 C戊', '', 'A戊 C戊', 'A戊 C戊', 'B戊,C戊'],
        [10, '古姿麟', 'PT034', 'FT', 'PT(物治)', 40, '', '', '', 'B甲', 'B甲', ''],
        [11, '簡廷宇', 'PT048', 'FT', 'PT(物治)', 40, '', '', '', 'B甲', '', ''],
        [12, '何沛錡', 'PT049', 'FT', 'PT(物治)', 40, '', 'C乙', 'B乙', 'C丙', 'A丙', 'C丙'],
        [13, '戴幸儀', 'OTP020', 'PT', 'OT(職治)', 40, '', '', '', '', '', ''],
        [14, '徐麗姿', 'PTP123', 'PT', 'PT(物治)', 0, '', '', '', '', '', ''],
        [15, '伍庭瑩', 'PTP125', 'PT', 'PT(物治)', 0, '', '', '', '', '', ''],
        [16, '朗振崴', 'PTP126', 'PT', 'PT(物治)', 0, '', '', 'A甲', '', '', ''],
        [17, '康宜姍', 'PTP114', 'PT', 'PT(物治)', 0, '', '', '', '', '', ''],
        [18, '蔡宗霖', 'PTP1127', 'PT', 'PT(物治)', 0, '', '', '', '', '', ''],
        [19, '馬奕凱', 'PTA003', 'FT', 'PT(物治)', 40, '', 'A甲,C戊', 'A戊,B戊', 'B戊,C戊', 'A戊,C戊', 'A戊,B戊'],
        [20, '林玉晴', 'PT003', 'FT', 'PT(物治)', 40, '', 'A戊,B戊', 'B戊,C戊', 'A戊,C戊', 'A戊,B戊', 'A甲,C戊']
    ]

    for row in real_staff_data: ws2.append(row)
    for cell in ws2[1]: cell.font = font_header; cell.fill = fill_header; cell.alignment = center_align
    
    dv_id = DataValidation(type="list", formula1='"FT,PT"', allow_blank=True); ws2.add_data_validation(dv_id); dv_id.add('D2:D100')
    dv_role = DataValidation(type="list", formula1='"PT(物治),OT(職治)"', allow_blank=True); ws2.add_data_validation(dv_role); dv_role.add('E2:E100')

    # Sheet 3: 例外請假
    ws3 = wb.create_sheet("3_例外請假")
    ws3.append(['姓名', '日期 (YYYY/MM/DD)', '時段', '類型 (下拉)', '備註'])
    for cell in ws3[1]: cell.font = font_header; cell.fill = fill_header
    dv_type = DataValidation(type="list", formula1='"OFF,ON"', allow_blank=True); ws3.add_data_validation(dv_type); dv_type.add('D2:D200')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# ⚙️ 第二部分：排班引擎邏輯 (V7.3)
# ==========================================
FIXED_LOCATIONS = ['甲', '乙', '丙']
DYNAMIC_LOCATIONS = ['丁', '戊']
ALL_LOCATIONS = FIXED_LOCATIONS + DYNAMIC_LOCATIONS
ROLE_PT = 'PT(物治)'
ROLE_OT = 'OT(職治)'

def find_best_candidates(needed_count, available_staff, d_str, shift, loc, role_filter, staff_db, calendar, exceptions, schedule):
    if needed_count <= 0: return []
    candidates = []
    dt_obj = datetime.strptime(d_str, '%Y/%m/%d')
    wk_idx = dt_obj.weekday()
    
    for name, info in staff_db.items():
        if role_filter and info['role'] != role_filter: continue
        day_load = 0; is_working_this_shift = False
        for l_chk in ALL_LOCATIONS:
            for s_chk in schedule[d_str]: 
                for worker in schedule[d_str][s_chk][l_chk]:
                    if worker['name'] == name:
                        day_load += 1
                        if s_chk == shift: is_working_this_shift = True
        if day_load >= 2: continue
        if is_working_this_shift: continue
        
        exc_key = (name, d_str, shift)
        if exceptions.get(exc_key) == 'OFF': continue
        
        rule_str = info['fixed_rules'].get(wk_idx, "")
        if info['type'] == 'PT':
            is_in_rules = shift in rule_str
            is_on_call = (exceptions.get(exc_key) == 'ON')
            if not (is_in_rules or is_on_call): continue 

        score = 0
        if info['type'] == 'FT': score += 1000 
        score -= (info['assigned_count'] * 10) 
        doc_name = calendar[d_str]['doctors'].get(loc, "")
        pair_count = info['doctor_history'].get(doc_name, 0)
        score -= pair_count 
        candidates.append({'name': name, 'score': score, 'type': info['type'], 'role': info['role'], 'id': info['id']})
    
    candidates.sort(key=lambda x: x['score'], reverse=True)
    return candidates[:needed_count]

def run_scheduler_bytes(input_file):
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
    except:
        return None, "❌ 無法讀取 Excel 檔案，請確認格式正確。"

    ws1 = wb['1_行事曆與醫師']
    calendar = {}; daily_requirements = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        date_val, wk, shift, doc_d, doc_e = row[:5]
        req_d_pt, req_d_ot, req_e_pt, req_e_ot, status = row[5:10]
        if not date_val: continue
        if status == '休診': continue
        d_str = date_val.strftime('%Y/%m/%d') if isinstance(date_val, datetime) else str(date_val).split(' ')[0]
        if d_str not in calendar: calendar[d_str] = {'shifts': set(), 'doctors': {}}
        calendar[d_str]['shifts'].add(shift)
        calendar[d_str]['doctors']['丁'] = doc_d; calendar[d_str]['doctors']['戊'] = doc_e
        daily_requirements[(d_str, shift, '丁', ROLE_PT)] = req_d_pt or 0
        daily_requirements[(d_str, shift, '丁', ROLE_OT)] = req_d_ot or 0
        daily_requirements[(d_str, shift, '戊', ROLE_PT)] = req_e_pt or 0
        daily_requirements[(d_str, shift, '戊', ROLE_OT)] = req_e_ot or 0

    ws2 = wb['2_人員設定']
    staff_db = {}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue 
        name = str(row[1]).strip()
        emp_id = str(row[2]).strip() if row[2] else "NO_ID"
        fixed_rules = {}
        for i in range(5): 
            val = row[7+i]; fixed_rules[i] = str(val).strip() if val else ""
        staff_db[name] = {
            'id': emp_id, 'type': str(row[3]).strip(), 'role': str(row[4]).strip(),
            'target': row[5] if isinstance(row[5], (int, float)) else 0,
            'fixed_rules': fixed_rules, 'assigned_count': 0, 'doctor_history': {}
        }

    ws3 = wb['3_例外請假']
    exceptions = {}
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]: continue
        e_d_str = row[1].strftime('%Y/%m/%d') if isinstance(row[1], datetime) else str(row[1]).split(' ')[0]
        exceptions[(str(row[0]).strip(), e_d_str, str(row[2]).strip())] = row[3]

    schedule = {}; sorted_dates = sorted(calendar.keys())
    for d_str in sorted_dates:
        schedule[d_str] = {'A':{}, 'B':{}, 'C':{}}
        for loc in ALL_LOCATIONS: schedule[d_str]['A'][loc] = []; schedule[d_str]['B'][loc] = []; schedule[d_str]['C'][loc] = []
            
    for d_str in sorted_dates:
        wk_idx = datetime.strptime(d_str, '%Y/%m/%d').weekday()
        if wk_idx > 4: continue 
        for name, info in staff_db.items():
            rule_str = info['fixed_rules'].get(wk_idx, "")
            for part in rule_str.replace('，', ',').split(','):
                clean_part = part.strip().replace('(', '').replace(')', '').replace(' ', '')
                match = re.match(r"([ABC])([甲乙丙丁戊])?", clean_part)
                if match:
                    s_code, l_code = match.groups()
                    if exceptions.get((name, d_str, s_code)) == 'OFF': continue
                    if l_code:
                        schedule[d_str][s_code][l_code].append({'name': name, 'type': info['type'], 'role': info['role'], 'is_fixed': True, 'id': info['id']})
                        staff_db[name]['assigned_count'] += 1

    for d_str in sorted_dates:
        for shift in sorted(list(calendar[d_str]['shifts'])):
            for loc in DYNAMIC_LOCATIONS:
                curr = schedule[d_str][shift][loc]
                needed_ot = daily_requirements.get((d_str, shift, loc, ROLE_OT), 0) - sum(1 for s in curr if s['role'] == ROLE_OT)
                if needed_ot > 0:
                    for p in find_best_candidates(needed_ot, staff_db, d_str, shift, loc, ROLE_OT, staff_db, calendar, exceptions, schedule):
                        schedule[d_str][shift][loc].append({'name': p['name'], 'type': p['type'], 'role': p['role'], 'is_fixed': False, 'id': p['id']})
                        staff_db[p['name']]['assigned_count'] += 1; needed_ot -= 1
                
                total_target = daily_requirements.get((d_str, shift, loc, ROLE_OT), 0) + daily_requirements.get((d_str, shift, loc, ROLE_PT), 0)
                final_needed = total_target - len(schedule[d_str][shift][loc])
                if final_needed > 0:
                    for p in find_best_candidates(final_needed, staff_db, d_str, shift, loc, ROLE_PT, staff_db, calendar, exceptions, schedule):
                        schedule[d_str][shift][loc].append({'name': p['name'], 'type': p['type'], 'role': p['role'], 'is_fixed': False, 'id': p['id']})
                        staff_db[p['name']]['assigned_count'] += 1; final_needed -= 1
                    if final_needed > 0:
                        for p in find_best_candidates(final_needed, staff_db, d_str, shift, loc, ROLE_OT, staff_db, calendar, exceptions, schedule):
                            if p['type'] == 'FT':
                                schedule[d_str][shift][loc].append({'name': p['name'], 'type': p['type'], 'role': p['role'], 'is_fixed': False, 'id': p['id']})
                                staff_db[p['name']]['assigned_count'] += 1; final_needed -= 1

    wb_out = Workbook()
    ws_dash = wb_out.active; ws_dash.title = "互動排班表"
    ws_raw = wb_out.create_sheet("原始運算底稿")
    
    fill_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') 
    fill_shifts = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid') 
    fill_gray  = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid') 
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ["姓名", "目標", "實際", "狀態", "A數", "B數", "C數", "AB天", "BC天", "AC天", "ABC天", "全休"]
    for idx, h in enumerate(headers, 1):
        cell = ws_dash.cell(6, idx, h); cell.fill = fill_green; cell.border = thin_border; cell.alignment = center_align
    
    staff_list = sorted([{'name':k, **v} for k,v in staff_db.items()], key=lambda x: str(x['id']))
    staff_row_map = {}
    row_offset = 7
    for i, info in enumerate(staff_list):
        r = row_offset + i
        staff_row_map[info['name']] = r
        ws_dash.cell(r, 1, info['name']).alignment = center_align
        ws_dash.cell(r, 2, info['target'] if info['type']=='FT' else "-").alignment = center_align
        c_cell, b_cell = f"C{r}", f"B{r}"
        f_status = f'=IF({c_cell}>{b_cell}, "加班 +"&({c_cell}-{b_cell}), IF({c_cell}<{b_cell}, "欠班 "&({c_cell}-{b_cell}), "正常"))' if info['type']=='FT' else f'="PT總診數: "&{c_cell}'
        ws_dash.cell(r, 4, f_status).alignment = center_align
        for c in range(1, 13): ws_dash.cell(r, c).border = thin_border

    col_cursor = 13; col_map = {}; date_col_ranges = {}
    for d_str in sorted_dates:
        start_col = col_cursor
        dt_obj = datetime.strptime(d_str, '%Y/%m/%d')
        shifts = sorted(list(calendar[d_str]['shifts']))
        for shift in shifts:
            for loc in ALL_LOCATIONS:
                ws_dash.cell(5, col_cursor, shift).alignment = center_align
                ws_dash.cell(6, col_cursor, loc).alignment = center_align
                ws_dash.cell(6, col_cursor).fill = fill_shifts
                ws_dash.cell(6, col_cursor).border = thin_border
                col_map[(d_str, shift, loc)] = col_cursor
                col_cursor += 1
        end_col = col_cursor - 1
        date_col_ranges[d_str] = (start_col, end_col)
        ws_dash.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        ws_dash.cell(3, start_col, dt_obj.strftime('%m/%d')).alignment = center_align
        ws_dash.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
        ws_dash.cell(4, start_col, ['一','二','三','四','五','六','日'][dt_obj.weekday()]).alignment = center_align

    for d_str in sorted_dates:
        for shift in schedule[d_str]:
            for loc in schedule[d_str][shift]:
                for worker in schedule[d_str][shift][loc]:
                    nm = worker['name']
                    if nm in staff_row_map and (d_str, shift, loc) in col_map:
                        c = col_map[(d_str, shift, loc)]
                        r = staff_row_map[nm]
                        ws_dash.cell(r, c, "V").alignment = center_align; ws_dash.cell(r, c).border = thin_border

    MAT_START, MAT_END = 13, col_cursor - 1
    for i in range(len(staff_list)):
        r = row_offset + i
        rng = f"{get_column_letter(MAT_START)}{r}:{get_column_letter(MAT_END)}{r}"
        hdr = f"${get_column_letter(MAT_START)}$5:${get_column_letter(MAT_END)}$5"
        ws_dash.cell(r, 3, f'=COUNTIF({rng}, "V")').alignment = center_align
        ws_dash.cell(r, 5, f'=COUNTIFS({hdr}, "A", {rng}, "V")').alignment = center_align
        ws_dash.cell(r, 6, f'=COUNTIFS({hdr}, "B", {rng}, "V")').alignment = center_align
        ws_dash.cell(r, 7, f'=COUNTIFS({hdr}, "C", {rng}, "V")').alignment = center_align

    dv = DataValidation(type="list", formula1='"V,休, "', allow_blank=True)
    ws_dash.add_data_validation(dv)
    dv.add(f"{get_column_letter(MAT_START)}7:{get_column_letter(MAT_END)}{row_offset + len(staff_list) - 1}")
    ws_dash.freeze_panes = "M7"

    raw_data = []
    for d_str in sorted_dates:
        for shift in schedule[d_str]:
            for loc in schedule[d_str][shift]:
                for worker in schedule[d_str][shift][loc]:
                    raw_data.append({'日期': d_str, '時段': shift, '地點': loc, '姓名': worker['name'], '員工編號': worker['id']})
    
    df_raw = pd.DataFrame(raw_data)
    for r in dataframe_to_rows(df_raw, index=False, header=True): ws_raw.append(r)

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output, "排班成功！儀表板已生成。"

# ==========================================
# ⚙️ 第三部分：ERP 轉檔邏輯 (V10)
# ==========================================
def convert_erp_bytes(input_file):
    try:
        df_raw = pd.read_excel(input_file, sheet_name='原始運算底稿')
    except:
        return None, "❌ 找不到「原始運算底稿」，請確認上傳的是排班結果檔。"
    
    if '員工編號' not in df_raw.columns: return None, "❌ 底稿中缺少「員工編號」，請重新執行排班。"

    df_raw['日期'] = pd.to_datetime(df_raw['日期'])
    staff_schedule = {}
    all_dates = sorted(df_raw['日期'].unique())
    
    for _, row in df_raw.iterrows():
        emp_id = str(row['員工編號']).strip()
        name = str(row['姓名']).strip()
        if emp_id == 'nan' or not emp_id: emp_id = "NO_ID"
        
        if emp_id not in staff_schedule: staff_schedule[emp_id] = {'name': name, 'data': {}}
        d_str = row['日期'].strftime('%Y/%m/%d')
        if d_str not in staff_schedule[emp_id]['data']: staff_schedule[emp_id]['data'][d_str] = []
        staff_schedule[emp_id]['data'][d_str].append({'shift': row['時段'], 'loc': row['地點']})
        
    wb_out = Workbook()
    ws_out = wb_out.active; ws_out.title = "ERP導入"
    
    color_header = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    color_id = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin = Side(style='thin', color="000000"); thick = Side(style='thick', color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thin, right=thin, top=thin, bottom=thick)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_out.merge_cells("A1:A2"); ws_out.merge_cells("B1:B2")
    ws_out.cell(1,1,"員工編號"); ws_out.cell(1,2,"姓名"); ws_out.cell(1,3,"星期"); ws_out.cell(2,3,"日期")
    
    weekday_map = {0:'一', 1:'二', 2:'三', 3:'四', 4:'五', 5:'六', 6:'日'}
    for i, dt in enumerate(all_dates):
        c = 4+i
        ws_out.cell(1,c, weekday_map[dt.weekday()])
        ws_out.cell(2,c, f"{dt.month}/{dt.day}")
        
    for r in [1,2]:
        for c in range(1, 4+len(all_dates)):
            cell = ws_out.cell(r,c); cell.fill = color_header; cell.alignment = center; cell.border = border_all; cell.font = Font(bold=True)
            
    curr_r = 3
    for emp_id in sorted(staff_schedule.keys()):
        data = staff_schedule[emp_id]
        ws_out.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r+2, end_column=1)
        ws_out.merge_cells(start_row=curr_r, start_column=2, end_row=curr_r+2, end_column=2)
        ws_out.cell(curr_r, 1, emp_id); ws_out.cell(curr_r, 2, data['name'])
        ws_out.cell(curr_r, 3, "班別排班"); ws_out.cell(curr_r+1, 3, "地點"); ws_out.cell(curr_r+2, 3, "備註")
        
        for i, dt in enumerate(all_dates):
            d_str = dt.strftime('%Y/%m/%d')
            c = 4+i
            if d_str in data['data']:
                items = sorted(data['data'][d_str], key=lambda x: {'A':1,'B':2,'C':3}.get(x['shift'], 9))
                ws_out.cell(curr_r, c, ",\n".join([x['shift'] for x in items]))
                ws_out.cell(curr_r+1, c, ",\n".join([x['loc'] for x in items]))
            ws_out.cell(curr_r+2, c, "")
            
        for r_idx in range(curr_r, curr_r+3):
            is_last = (r_idx == curr_r+2)
            bd = border_thick if is_last else border_all
            for c_idx in range(1, 4+len(all_dates)):
                cell = ws_out.cell(r_idx, c_idx); cell.border = bd; cell.alignment = center
                if c_idx==1: cell.fill = color_id
        
        curr_r += 3
        
    ws_out.column_dimensions['A'].width = 15; ws_out.column_dimensions['B'].width = 15; ws_out.column_dimensions['C'].width = 12
    for c in range(4, 4+len(all_dates)): ws_out.column_dimensions[get_column_letter(c)].width = 6

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)
    return output, "ERP 轉檔成功！"

# ==========================================
# 📱 網頁介面 (Streamlit UI)
# ==========================================
st.set_page_config(page_title="晉安毅安復健治療師智慧排班系統", layout="wide", page_icon="🏥")

# CSS 美化標題與區塊
st.markdown("""
    <style>
    .main-title {
        font-size: 36px;
        font-weight: bold;
        color: #2F75B5;
        text-align: center;
        margin-bottom: 20px;
    }
    .sub-title {
        font-size: 20px;
        color: #555;
        text-align: center;
        margin-bottom: 30px;
    }
    </style>
    <div class="main-title">🏥 晉安毅安復健治療師智慧排班系統</div>
    <div class="sub-title">自動化排班流程：產生模板 ➡️ 執行排班 ➡️ 轉檔 ERP</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["1️⃣ 第一步：產生輸入表", "2️⃣ 第二步：執行排班", "3️⃣ 第三步：轉檔 ERP"])

with tab1:
    st.header("產生空白輸入表 (模板)")
    st.info("請選擇要排班的年份與月份，下載後的 Excel 已包含預填的醫師班表與真實人員名單。")
    col1, col2 = st.columns(2)
    with col1: year = st.number_input("年份", min_value=2024, max_value=2030, value=2026)
    with col2: month = st.number_input("月份", min_value=1, max_value=12, value=3)
    
    if st.button("🚀 產生輸入表 (模板)", type="primary"):
        file_bytes = generate_template_bytes(year, month)
        st.success(f"✅ 已產生 {year}年{month}月 的輸入表！")
        st.download_button(
            label="📥 下載 Excel 模板",
            data=file_bytes,
            file_name=f"【復健部輸入表】{year}年{month}月_真實資料版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with tab2:
    st.header("執行排班運算")
    st.info("請上傳填寫好的輸入表，系統將自動進行瀑布流排班，並產出互動式儀表板。")
    uploaded_file = st.file_uploader("上傳 Step 1 的 Excel 檔案", type=['xlsx'])
    
    if uploaded_file is not None:
        if st.button("⚡ 開始排班", type="primary"):
            with st.spinner('正在進行複雜排班運算 (A/B/C 三診 + 瀑布流 + 跨界支援)...'):
                result_bytes, msg = run_scheduler_bytes(uploaded_file)
            
            if result_bytes:
                st.balloons()
                st.success(f"✅ {msg}")
                st.download_button(
                    label="📥 下載排班結果 (含儀表板)",
                    data=result_bytes,
                    file_name="【復健部排班結果】V7_3_儀表板版.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error(msg)

with tab3:
    st.header("轉出 ERP 格式")
    st.info("請上傳 Step 2 的排班結果，系統將自動轉換為符合 ERP 導入標準的綠色表格。")
    result_file = st.file_uploader("上傳 Step 2 的 Excel 檔案", type=['xlsx'], key="erp")
    
    if result_file is not None:
        if st.button("🔄 轉換為 ERP 格式", type="primary"):
            erp_bytes, msg = convert_erp_bytes(result_file)
            if erp_bytes:
                st.balloons()
                st.success(f"✅ {msg}")
                st.download_button(
                    label="📥 下載 ERP 導入檔",
                    data=erp_bytes,
                    file_name="ERP導入檔_復健部_V10_完美版.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error(msg)