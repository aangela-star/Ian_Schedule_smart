import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import io
import re
import random
import numpy as np


# ==========================================
# 🔒 安全守門員：登入檢查系統
# ==========================================
def check_password():
    """如果使用者輸入正確密碼，回傳 True，否則回傳 False"""

    def password_entered():
        """檢查使用者輸入的密碼是否與 secrets 中的設定相符"""
        if st.session_state["password"] == st.secrets["LOGIN_PASSWORD"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # 驗證後刪除輸入框的暫存，保持乾淨
        else:
            st.session_state["password_correct"] = False

    # 初始化 session state
    if "password_correct" not in st.session_state:
        # 第一次進入，顯示輸入框
        st.text_input(
            "請輸入系統密碼 / Password", type="password", on_change=password_entered, key="password"
        )
        return False
    
    elif not st.session_state["password_correct"]:
        # 密碼錯誤，再次顯示輸入框
        st.text_input(
            "❌ 密碼錯誤，請重試 / Password", type="password", on_change=password_entered, key="password"
        )
        return False
    
    else:
        # 密碼正確
        return True

# 🚨 執行檢查：如果沒通過，程式就停在這裡 (st.stop)
if not check_password():
    st.stop()

# ==========================================
# 👇 只有登入成功後，才會執行下面的程式碼
# ==========================================


# ==========================================
# ⚙️ 第一部分：產生模板 (V5 + 真實名單)
# ==========================================
def generate_nurse_template_bytes(year, month):
    wb = Workbook()
    
    # 紫色系樣式
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid") # 紫色
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Sheet 0
    ws0 = wb.active; ws0.title = "0_全域控制台"
    ws0.append(['項目', '數值', '說明'])
    ws0.append(['年份', year, '設定排班年份'])
    ws0.append(['月份', month, '設定排班月份'])
    for cell in ws0[1]: cell.font = font_header; cell.fill = fill_header

    # Sheet 1: 行事曆
    ws1 = wb.create_sheet("1_醫師班表與營業日")
    dates = pd.date_range(start=f'{year}-{month}-01', end=f'{year}-{month}-{pd.Period(f"{year}-{month}").days_in_month}')
    weekday_map = {0:'一', 1:'二', 2:'三', 3:'四', 4:'五', 5:'六', 6:'日'}
    ws1.append(['日期', '星期', '時段', '甲院_醫師', '乙院_醫師', '營業狀態'])
    
    row_count = 1
    for d in dates:
        if d.weekday() >= 6: continue # 週日休
        d_str = d.strftime('%Y/%m/%d')
        wk = weekday_map[d.weekday()]
        status = '營業'
        for shift in ['A', 'B', 'C']:
            # 簡單預填，讓使用者自己改
            doc_a = '劉醫師' if shift != 'C' else '莊醫師'
            doc_b = '王醫師' if shift != 'B' else '薛醫師'
            ws1.append([d_str, wk, shift, doc_a, doc_b, status])
            row_count += 1
            
    for cell in ws1[1]: cell.font = font_header; cell.fill = fill_header; cell.alignment = center_align
    if row_count > 1:
        dv = DataValidation(type="list", formula1='"營業,休診"', allow_blank=False)
        ws1.add_data_validation(dv); dv.add(f'F2:F{row_count}')

    # Sheet 2: 人員設定 (★修正：預填真實名單★)
    ws2 = wb.create_sheet("2_人員設定")
    headers2 = ['序號', '姓名', '員工編號', '身分 (下拉)', '職能 (下拉)', '本月個人目標 (數字)', '備註', '週一 (固定)', '週二 (固定)', '週三 (固定)', '週四 (固定)', '週五 (固定)', '週六 (固定)']
    ws2.append(headers2)
    
    # 真實人員資料 (依據提供的圖片)
    # [序號, 姓名, 員編, 身分, 職能, 目標(預設40/0), 備註, 週一~週六固定班]
    real_staff_data = [
        [1, '品', 'NS014', 'FT', 'Nurse', 40, '', '', '', '', '', '', ''],
        [2, '智', 'NS028', 'FT', 'Nurse', 40, '', '', '', '', '', '', ''],
        [3, '廖', 'NS031', 'FT', 'Nurse', 40, '', '', '', '', '', '', ''],
        [4, '淑', 'FD043', 'FT', 'Admin', 40, '', '', '', '', '', '', ''],
        [5, '喬', 'FD021', 'FT', 'Admin', 40, '', '', '', '', '', '', ''],
        [6, '淇', 'FD032', 'FT', 'Admin', 40, '', '', '', '', '', '', ''],
        [7, '芯', 'FD054', 'PT', 'Admin', 0,  '', '', '', '', '', '', ''],
        [8, '圩', 'FD053', 'PT', 'Admin', 0,  '', '', '', '', '', '', '']
    ]

    for row in real_staff_data:
        ws2.append(row)
    
    for cell in ws2[1]: cell.font = font_header; cell.fill = fill_header; cell.alignment = center_align
    dv_id = DataValidation(type="list", formula1='"FT,PT"', allow_blank=True); ws2.add_data_validation(dv_id); dv_id.add('D2:D100')
    dv_role = DataValidation(type="list", formula1='"Nurse,Admin"', allow_blank=True); ws2.add_data_validation(dv_role); dv_role.add('E2:E100')

    # Sheet 3
    ws3 = wb.create_sheet("3_例外請假")
    ws3.append(['姓名', '日期 (YYYY/MM/DD)', '時段 (下拉)', '類型 (下拉)', '備註'])
    for cell in ws3[1]: cell.font = font_header; cell.fill = fill_header
    dv_type = DataValidation(type="list", formula1='"OFF,ON,PT_OK"', allow_blank=True); ws3.add_data_validation(dv_type); dv_type.add('D2:D200')

    # Sheet 4
    ws4 = wb.create_sheet("4_醫師人力規則")
    ws4.append(['醫師姓名 (關鍵字)', '需配置人力'])
    ws4.append(['劉醫師', 3]); ws4.append(['莊醫師', 2]); ws4.append(['薛醫師', 2]); ws4.append(['預設值', 2])
    for cell in ws4[1]: cell.font = font_header; cell.fill = fill_header

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# ⚙️ 第二部分：排班引擎 (V10.13 邏輯移植)
# ==========================================
class ClinicSchedulerNurse:
    def __init__(self, input_file):
        self.input_file = input_file
        self.schedule_log_matrix = []
        self.schedule_log_v8 = []
        self.staff_targets = {}
        self.off_lookup_map = {} 
        self.on_lookup_map = {}  
        self.doctor_load_map = {}
        
    def load_data(self):
        try:
            self.df_calendar = pd.read_excel(self.input_file, sheet_name='1_醫師班表與營業日')
            self.df_calendar['日期'] = pd.to_datetime(self.df_calendar['日期']).dt.normalize()
            
            self.df_staff = pd.read_excel(self.input_file, sheet_name='2_人員設定')
            self.df_staff['姓名'] = self.df_staff['姓名'].astype(str).str.replace(' ', '')
            self.df_staff['員工編號'] = self.df_staff['員工編號'].astype(str).str.strip().replace('nan', 'NO_ID')
            self.staff_targets = dict(zip(self.df_staff['姓名'], self.df_staff['本月個人目標 (數字)'].fillna(0)))
            
            self.df_wishes = pd.read_excel(self.input_file, sheet_name='3_例外請假')
            self.df_wishes['日期'] = pd.to_datetime(self.df_wishes['日期 (YYYY/MM/DD)']).dt.normalize()
            
            for _, row in self.df_wishes.iterrows():
                name = str(row['姓名']).strip()
                date_str = row['日期'].strftime('%Y/%m/%d')
                w_type = row['類型 (下拉)']
                shift = str(row['時段 (下拉)']).upper() if pd.notna(row['時段 (下拉)']) else "ABC"
                key = (name, date_str)
                if w_type == 'OFF': self.off_lookup_map[key] = self.off_lookup_map.get(key, "") + shift
                elif w_type in ['ON', 'PT_OK']: self.on_lookup_map[key] = self.on_lookup_map.get(key, "") + shift

            df_rules = pd.read_excel(self.input_file, sheet_name='4_醫師人力規則')
            self.doctor_load_map = dict(zip(df_rules['醫師姓名 (關鍵字)'], df_rules['需配置人力']))
            return True, "資料讀取成功"
        except Exception as e:
            return False, f"讀取失敗: {e}"

    def get_required_staff_count(self, doctor_name):
        doc_str = str(doctor_name).strip()
        if doc_str in ['nan', 'None', '', '無']: return 0
        for k, v in self.doctor_load_map.items():
            if k in doc_str: return int(v)
        return self.doctor_load_map.get('預設值', 2)

    def is_available(self, staff_row, date_ts, shift):
        name = staff_row['姓名']
        d_str = date_ts.strftime('%Y/%m/%d')
        
        # OFF Check
        if (name, d_str) in self.off_lookup_map:
            if shift in self.off_lookup_map[(name, d_str)]: return False
            
        # ON Check
        if (name, d_str) in self.on_lookup_map:
            if shift in self.on_lookup_map[(name, d_str)]: return True
            
        # Fixed Rule Check
        wk_map = {0:'週一', 1:'週二', 2:'週三', 3:'週四', 4:'週五', 5:'週六', 6:'週日'}
        col_name = f"{wk_map[date_ts.weekday()]} (固定)"
        rule = str(staff_row.get(col_name, '')).upper()
        
        if staff_row['身分 (下拉)'] == 'PT':
            if rule in ['NAN', '', '0']: return False
            return shift in rule
        else: # FT
            if rule not in ['NAN', '', '0']: return shift in rule
            return True 

    def run(self):
        dates = sorted(self.df_calendar['日期'].unique())
        staff_counts = {name: 0 for name in self.df_staff['姓名']}
        
        nurses = self.df_staff[(self.df_staff['身分 (下拉)']=='FT') & (self.df_staff['職能 (下拉)']=='Nurse')]
        admins = self.df_staff[(self.df_staff['身分 (下拉)']=='FT') & (self.df_staff['職能 (下拉)']=='Admin')]
        pts = self.df_staff[self.df_staff['身分 (下拉)']=='PT']
        
        nurse_names = nurses['姓名'].tolist()
        admin_names = admins['姓名'].tolist()
        n_idx = 0; a_idx = 0
        
        for d in dates:
            day_data = self.df_calendar[self.df_calendar['日期'] == d]
            
            # 護理師輪替邏輯 (N1/N2/N3)
            today_nurse_ptr = {}
            if len(nurse_names) >= 3:
                n1, n2, n3 = nurse_names[n_idx%len(nurse_names)], nurse_names[(n_idx+1)%len(nurse_names)], nurse_names[(n_idx+2)%len(nurse_names)]
                today_nurse_ptr = {n1:['A','B'], n2:['B','C'], n3:['A','C']}
                n_idx += 1
            
            # 行政輪替
            curr_admins = admin_names[a_idx%len(admin_names):] + admin_names[:a_idx%len(admin_names)]
            a_idx += 1
            
            for shift in ['A', 'B', 'C']:
                row = day_data[day_data['時段'] == shift]
                if row.empty or row.iloc[0]['營業狀態'] != '營業': continue
                row = row.iloc[0]
                
                req_a = self.get_required_staff_count(row['甲院_醫師'])
                req_b = self.get_required_staff_count(row['乙院_醫師'])
                
                assigned_a = []; assigned_b = []
                
                # 建立候選人池
                pool_n = []; pool_a = []; pool_p = []
                
                # 1. Nurses
                for _, n in nurses.iterrows():
                    nm = n['姓名']
                    if self.is_available(n, d, shift):
                        # 優先權：輪值 > 欠班 > 其他
                        score = 100
                        if nm in today_nurse_ptr and shift in today_nurse_ptr[nm]: score += 500
                        if staff_counts[nm] < self.staff_targets.get(nm,0): score += 50
                        pool_n.append({'name': nm, 'score': score, 'type': 'N', 'id': n['員工編號']})
                pool_n.sort(key=lambda x: x['score'], reverse=True)
                
                # 2. Admins
                for _, a in admins.iterrows():
                    nm = a['姓名']
                    if self.is_available(a, d, shift):
                        score = 50
                        if nm == curr_admins[0]: score += 100 # 今日優先
                        if staff_counts[nm] < self.staff_targets.get(nm,0): score += 50
                        pool_a.append({'name': nm, 'score': score, 'type': 'A', 'id': a['員工編號']})
                pool_a.sort(key=lambda x: x['score'], reverse=True)
                
                # 3. PTs
                for _, p in pts.iterrows():
                    if self.is_available(p, d, shift):
                        pool_p.append({'name': p['姓名'], 'score': 10, 'type': 'PT', 'id': p['員工編號']})
                
                # 填補邏輯：優先填 N -> A -> PT
                def get_staff():
                    if pool_n: return pool_n.pop(0)
                    if pool_a: return pool_a.pop(0)
                    if pool_p: return pool_p.pop(0)
                    return None
                
                while len(assigned_a) < req_a:
                    s = get_staff()
                    if s: assigned_a.append(s); staff_counts[s['name']] += 1
                    else: break
                    
                while len(assigned_b) < req_b:
                    s = get_staff()
                    if s: assigned_b.append(s); staff_counts[s['name']] += 1
                    else: break
                
                # 紀錄結果
                for s in assigned_a:
                    self.schedule_log_matrix.append({'日期': d, '時段': shift, '地點': '甲', '姓名': s['name'], '員工編號': s['id']})
                for s in assigned_b:
                    self.schedule_log_matrix.append({'日期': d, '時段': shift, '地點': '乙', '姓名': s['name'], '員工編號': s['id']})

        return self.generate_excel()

    def generate_excel(self):
        wb = Workbook()
        ws = wb.active; ws.title = "互動排班表"
        ws_raw = wb.create_sheet("原始運算底稿")
        
        # 紫色系
        fill_purple = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid') # 淺紫
        fill_dark_p = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid') # 深紫
        font_white = Font(color="FFFFFF", bold=True)
        thin = Side(style='thin'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center')
        
        # Dashboard Headers
        headers = ["姓名", "目標", "實際", "狀態", "A數", "B數", "C數", "AB天", "BC天", "AC天", "ABC天", "全休"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(6, i, h); cell.fill = fill_dark_p; cell.font = font_white; cell.border = border; cell.alignment = center
            
        # Staff Rows
        staff_list = self.df_staff.to_dict('records')
        row_map = {}
        for i, s in enumerate(staff_list):
            r = 7 + i
            row_map[s['姓名']] = r
            ws.cell(r, 1, s['姓名']).alignment = center
            ws.cell(r, 2, s['本月個人目標 (數字)']).alignment = center
            # 公式
            c_cell, b_cell = f"C{r}", f"B{r}"
            f_stat = f'=IF({c_cell}>{b_cell}, "加班 +"&({c_cell}-{b_cell}), IF({c_cell}<{b_cell}, "欠班 "&({c_cell}-{b_cell}), "正常"))'
            if s['身分 (下拉)'] == 'PT': f_stat = f'="PT: "&{c_cell}'
            ws.cell(r, 4, f_stat).alignment = center
            for c in range(1, 13): ws.cell(r, c).border = border

        # Matrix
        col = 13; col_map = {}; dates = sorted(self.df_calendar['日期'].unique())
        for d in dates:
            start_c = col
            dt_obj = d.to_pydatetime()
            for shift in ['A', 'B', 'C']:
                for loc in ['甲', '乙']:
                    ws.cell(5, col, shift).alignment = center
                    ws.cell(6, col, loc).alignment = center; ws.cell(6, col).fill = fill_purple; ws.cell(6, col).border = border
                    col_map[(d, shift, loc)] = col
                    col += 1
            end_c = col - 1
            ws.merge_cells(start_row=3, start_column=start_c, end_row=3, end_column=end_c)
            ws.cell(3, start_c, dt_obj.strftime('%m/%d')).alignment = center
            
        # Fill Data
        for rec in self.schedule_log_matrix:
            key = (rec['日期'], rec['時段'], rec['地點'])
            nm = rec['姓名']
            if key in col_map and nm in row_map:
                r, c = row_map[nm], col_map[key]
                ws.cell(r, c, "V").alignment = center; ws.cell(r, c).border = border
                
        # Fill OFF
        for (nm, d_str), shifts in self.off_lookup_map.items():
            if nm in row_map:
                r = row_map[nm]
                pass 

        # Formulas
        M_S, M_E = 13, col - 1
        for i in range(len(staff_list)):
            r = 7 + i
            rng = f"{get_column_letter(M_S)}{r}:{get_column_letter(M_E)}{r}"
            hdr = f"${get_column_letter(M_S)}$5:${get_column_letter(M_E)}$5"
            ws.cell(r, 3, f'=COUNTIF({rng}, "V")').alignment = center
            ws.cell(r, 5, f'=COUNTIFS({hdr}, "A", {rng}, "V")').alignment = center
            ws.cell(r, 6, f'=COUNTIFS({hdr}, "B", {rng}, "V")').alignment = center
            ws.cell(r, 7, f'=COUNTIFS({hdr}, "C", {rng}, "V")').alignment = center

        ws.freeze_panes = "M7"
        
        # Raw Data
        df_raw = pd.DataFrame(self.schedule_log_matrix)
        for row in dataframe_to_rows(df_raw, index=False, header=True): ws_raw.append(row)
        
        output = io.BytesIO()
        wb.save(output); output.seek(0)
        return output

def run_nurse_scheduler(input_file):
    scheduler = ClinicSchedulerNurse(input_file)
    success, msg = scheduler.load_data()
    if not success: return None, msg
    return scheduler.run(), "排班成功"

# ==========================================
# ⚙️ 第三部分：ERP 轉檔 (V10)
# ==========================================
def convert_nurse_erp(input_file):
    try:
        df_raw = pd.read_excel(input_file, sheet_name='原始運算底稿')
    except: return None, "❌ 找不到底稿"
    
    if '員工編號' not in df_raw.columns: return None, "❌ 缺少員編"
    
    df_raw['日期'] = pd.to_datetime(df_raw['日期'])
    staff_data = {}
    dates = sorted(df_raw['日期'].unique())
    
    for _, row in df_raw.iterrows():
        eid = str(row['員工編號']).strip()
        nm = str(row['姓名']).strip()
        if eid == 'nan': eid = "NO_ID"
        if eid not in staff_data: staff_data[eid] = {'name': nm, 'data': {}}
        d_str = row['日期'].strftime('%Y/%m/%d')
        if d_str not in staff_data[eid]['data']: staff_data[eid]['data'][d_str] = []
        staff_data[eid]['data'][d_str].append({'s': row['時段'], 'l': row['地點']})
        
    wb = Workbook(); ws = wb.active; ws.title = "ERP導入"
    fill_h = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid") # 紫色
    fill_id = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid") # 淺紫
    font_w = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.merge_cells("A1:A2"); ws.merge_cells("B1:B2")
    ws.cell(1,1,"員工編號"); ws.cell(1,2,"姓名"); ws.cell(1,3,"星期"); ws.cell(2,3,"日期")
    
    wk_map = {0:'一', 1:'二', 2:'三', 3:'四', 4:'五', 5:'六', 6:'日'}
    for i, d in enumerate(dates):
        c = 4+i
        ws.cell(1, c, wk_map[d.weekday()])
        ws.cell(2, c, f"{d.month}/{d.day}")
        
    for r in [1,2]:
        for c in range(1, 4+len(dates)):
            cell = ws.cell(r,c); cell.fill = fill_h; cell.font = font_w; cell.border = border; cell.alignment = center
            
    curr_r = 3
    for eid in sorted(staff_data.keys()):
        data = staff_data[eid]
        ws.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r+2, end_column=1)
        ws.merge_cells(start_row=curr_r, start_column=2, end_row=curr_r+2, end_column=2)
        ws.cell(curr_r, 1, eid); ws.cell(curr_r, 2, data['name'])
        ws.cell(curr_r, 3, "班別排班"); ws.cell(curr_r+1, 3, "地點"); ws.cell(curr_r+2, 3, "備註")
        
        for i, d in enumerate(dates):
            d_str = d.strftime('%Y/%m/%d')
            c = 4+i
            if d_str in data['data']:
                items = sorted(data['data'][d_str], key=lambda x: x['s'])
                ws.cell(curr_r, c, ",\n".join([x['s'] for x in items]))
                ws.cell(curr_r+1, c, ",\n".join([x['l'] for x in items]))
            ws.cell(curr_r+2, c, "")
            
        for r_idx in range(curr_r, curr_r+3):
            for c_idx in range(1, 4+len(dates)):
                cell = ws.cell(r_idx, c_idx); cell.border = border; cell.alignment = center
                if c_idx==1: cell.fill = fill_id
        curr_r += 3
        
    ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 12; ws.column_dimensions['C'].width = 12
    for c in range(4, 4+len(dates)): ws.column_dimensions[get_column_letter(c)].width = 6
    
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return output, "轉檔成功"

# ==========================================
# 📱 介面 (Purple Theme)
# ==========================================
st.set_page_config(page_title="晉安毅安護理師智慧排班系統", layout="wide", page_icon="💉")

st.markdown("""
    <style>
    .main-title { font-size: 36px; font-weight: bold; color: #7030A0; text-align: center; margin-bottom: 20px; }
    .sub-title { font-size: 20px; color: #555; text-align: center; margin-bottom: 30px; }
    </style>
    <div class="main-title">💉 晉安毅安護理師智慧排班系統</div>
    <div class="sub-title">自動化排班流程：產生模板 ➡️ 執行排班 ➡️ 轉檔 ERP</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["1️⃣ 產生模板", "2️⃣ 執行排班", "3️⃣ 轉檔 ERP"])

with tab1:
    st.header("產生空白輸入表 (模板)")
    st.info("請選擇年份與月份，下載後的 Excel 已包含預填的真實人員資料。")
    c1, c2 = st.columns(2)
    with c1: year = st.number_input("年份", 2024, 2030, 2026)
    with c2: month = st.number_input("月份", 1, 12, 2)
    if st.button("🚀 下載模板", type="primary"):
        st.download_button("📥 下載 Excel", generate_nurse_template_bytes(year, month), 
                           f"【護理師輸入表】{year}年{month}月_真實資料版.xlsx")

with tab2:
    st.header("執行排班")
    f = st.file_uploader("上傳輸入表", type=['xlsx'])
    if f and st.button("⚡ 開始排班", type="primary"):
        with st.spinner("正在進行護理師輪替排班..."):
            res, msg = run_nurse_scheduler(f)
            if res: st.success(msg); st.download_button("📥 下載結果", res, "【護理師排班結果】V10_儀表板版.xlsx")
            else: st.error(msg)

with tab3:
    st.header("轉出 ERP")
    f2 = st.file_uploader("上傳結果檔", type=['xlsx'], key='erp')
    if f2 and st.button("🔄 轉檔", type="primary"):
        res, msg = convert_nurse_erp(f2)
        if res: st.success(msg); st.download_button("📥 下載 ERP 檔", res, "ERP導入檔_護理師.xlsx")
        else: st.error(msg)